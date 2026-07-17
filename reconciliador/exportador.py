"""
exportador.py

Responsável pela geração do relatório Excel.
"""

from datetime import datetime
from io import BytesIO

import pandas as pd

from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side
)

from openpyxl.utils import (
    get_column_letter
)


# ============================================================
# CORES
# ============================================================

AZUL = "1F4E78"
VERDE = "70AD47"
VERMELHO = "C00000"
LARANJA = "ED7D31"
AMARELO = "FFD966"
CINZA = "D9E1F2"

BRANCO = "FFFFFF"
PRETO = "000000"


# ============================================================
# ESTILOS
# ============================================================

FILL_TITULO = PatternFill(
    "solid",
    fgColor=AZUL
)

FILL_SECAO = PatternFill(
    "solid",
    fgColor=CINZA
)

FILL_ALERTA = PatternFill(
    "solid",
    fgColor=AMARELO
)

FONTE_TITULO = Font(
    bold=True,
    color=BRANCO,
    size=12
)

FONTE_SECAO = Font(
    bold=True,
    size=11
)

FONTE_NORMAL = Font(
    size=10
)

BORDA = Border(

    left=Side(style="thin"),

    right=Side(style="thin"),

    top=Side(style="thin"),

    bottom=Side(style="thin")

)


# ============================================================
# FUNÇÕES AUXILIARES
# ============================================================

def aplicar_cabecalho(ws):

    for celula in ws[1]:

        celula.fill = FILL_TITULO

        celula.font = FONTE_TITULO

        celula.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        celula.border = BORDA


def congelar(ws):

    ws.freeze_panes = "A2"


def aplicar_filtro(ws):

    ws.auto_filter.ref = ws.dimensions


def ajustar_colunas(ws):

    for coluna in ws.columns:

        tamanho = 0

        letra = get_column_letter(
            coluna[0].column
        )

        for celula in coluna:

            try:

                tamanho = max(
                    tamanho,
                    len(str(celula.value))
                )

            except Exception:

                pass

        ws.column_dimensions[
            letra
        ].width = min(
            tamanho + 4,
            60
        )


def formatar_tabela(ws):

    aplicar_cabecalho(ws)

    congelar(ws)

    aplicar_filtro(ws)

    ajustar_colunas(ws)
def nome_bonito(nome):
    nome = nome.replace("_", " ")

    ajustes = {
        "numero nf": "Número NF",
        "cnpj": "CNPJ",
        "cpf": "CPF",
        "id": "ID",
        "nf": "NF",
    }

    return ajustes.get(nome.lower(), nome.title())


def renomear_colunas(df):
    df = df.copy()

    novos_nomes = {}

    for coluna in df.columns:

        if coluna == "_chave_reconciliacao":
            continue

        elif coluna == "_origem":
            novos_nomes[coluna] = "Planilha"

        elif coluna == "_linha_origem":
            novos_nomes[coluna] = "Linha"

        elif coluna.startswith("_raw_"):
            conceito = nome_bonito(coluna.replace("_raw_", ""))
            novos_nomes[coluna] = f"{conceito} (original)"

        elif coluna.startswith("_status_"):
            conceito = nome_bonito(coluna.replace("_status_", ""))
            novos_nomes[coluna] = f"{conceito} (status)"

        elif coluna.startswith("_mensagem_"):
            conceito = nome_bonito(coluna.replace("_mensagem_", ""))
            novos_nomes[coluna] = f"{conceito} (observação)"

        elif not coluna.startswith("_"):
            novos_nomes[coluna] = nome_bonito(coluna)

    df = df.drop(
        columns=["_chave_reconciliacao"],
        errors="ignore"
    )

    return df.rename(columns=novos_nomes)

# ============================================================
# DASHBOARD
# ============================================================

def escrever_dashboard(ws, resultado):

    ws["A1"] = "RECONCILIADOR DE PLANILHAS"
    ws["A1"].font = Font(
        bold=True,
        size=18,
        color=BRANCO
    )
    ws["A1"].fill = FILL_TITULO

    ws.merge_cells("A1:D1")

    ws["A3"] = "Data da comparação"
    ws["A3"].font = FONTE_SECAO

    ws["B3"] = datetime.now().strftime(
        "%d/%m/%Y %H:%M:%S"
    )

    ws["A5"] = "PLANILHAS"
    ws["A5"].font = FONTE_SECAO
    ws["A5"].fill = FILL_SECAO

    ws["A6"] = "Planilha A"
    ws["B6"] = resultado.resumo.nome_planilha_a

    ws["A7"] = "Planilha B"
    ws["B7"] = resultado.resumo.nome_planilha_b

    ws["A9"] = "RESUMO"
    ws["A9"].font = FONTE_SECAO
    ws["A9"].fill = FILL_SECAO

    indicadores = [

        ("Registros Planilha A",
         resultado.resumo.total_planilha_a),

        ("Registros Planilha B",
         resultado.resumo.total_planilha_b),

        ("Consistentes",
         resultado.resumo.consistentes),

        ("Exclusivos Planilha A",
         resultado.resumo.exclusivos_a),

        ("Exclusivos Planilha B",
         resultado.resumo.exclusivos_b),

        ("Inconsistentes",
         resultado.resumo.inconsistentes),

        ("Consolidado",
         resultado.resumo.consolidado)

    ]

    linha = 10

    for indicador, valor in indicadores:

        ws[f"A{linha}"] = indicador
        ws[f"B{linha}"] = valor

        linha += 1

    # ==========================================
    # Percentuais
    # ==========================================

    total = max(
        resultado.resumo.consolidado,
        1
    )

    ws["D5"] = "INDICADORES"
    ws["D5"].font = FONTE_SECAO
    ws["D5"].fill = FILL_SECAO

    metricas = [

        ("Consistentes",
         resultado.resumo.consistentes),

        ("Exclusivos A",
         resultado.resumo.exclusivos_a),

        ("Exclusivos B",
         resultado.resumo.exclusivos_b),

        ("Inconsistentes",
         resultado.resumo.inconsistentes)

    ]

    linha = 6

    for nome, valor in metricas:

        percentual = valor / total

        barras = "█" * int(percentual * 20)

        ws[f"D{linha}"] = nome

        ws[f"E{linha}"] = barras

        ws[f"F{linha}"] = f"{percentual:.1%}"

        linha += 1

    ajustar_colunas(ws)

def renomear_colunas(df):
    df = df.copy()

    novos_nomes = {}

    for coluna in df.columns:

        if coluna == "_chave_reconciliacao":
            continue

        elif coluna == "_origem":
            novos_nomes[coluna] = "Planilha"

        elif coluna == "_linha_origem":
            novos_nomes[coluna] = "Linha"

        elif coluna.startswith("_raw_"):
            conceito = coluna.replace("_raw_", "").replace("_", " ").title()
            novos_nomes[coluna] = f"{conceito} (original)"

        elif coluna.startswith("_status_"):
            conceito = coluna.replace("_status_", "").replace("_", " ").title()
            novos_nomes[coluna] = f"{conceito} (status)"

        elif coluna.startswith("_mensagem_"):
            conceito = coluna.replace("_mensagem_", "").replace("_", " ").title()
            novos_nomes[coluna] = f"{conceito} (observação)"

    df = df.drop(
        columns=["_chave_reconciliacao"],
        errors="ignore"
    )

    return df.rename(columns=novos_nomes)

# ============================================================
# ESCRITA DAS ABAS
# ============================================================

def escrever_dataframe(
    writer,
    nome_aba,
    dataframe,
    destacar_colunas=None
):

    dataframe = renomear_colunas(dataframe)

    dataframe.to_excel(
        writer,
        sheet_name=nome_aba,
        index=False
    )

    ws = writer.sheets[nome_aba]

    formatar_tabela(ws)

    if destacar_colunas:

        cabecalhos = {}

        for coluna in range(1, ws.max_column + 1):

            cabecalhos[
                ws.cell(
                    row=1,
                    column=coluna
                ).value
            ] = coluna

        for nome in destacar_colunas:

            if nome not in cabecalhos:
                continue

            coluna = cabecalhos[nome]

            for linha in range(
                2,
                ws.max_row + 1
            ):

                celula = ws.cell(
                    row=linha,
                    column=coluna
                )

                celula.fill = FILL_ALERTA


# ============================================================
# DASHBOARD
# ============================================================

def escrever_aba_dashboard(
    writer,
    resultado
):

    ws = writer.book.create_sheet(
        "Dashboard",
        0
    )

    escrever_dashboard(
        ws,
        resultado
    )


# ============================================================
# PLANILHAS
# ============================================================

def escrever_planilhas(
    writer,
    resultado
):

    escrever_dataframe(
        writer,
        "Consistentes",
        resultado.consistentes
    )

    escrever_dataframe(
        writer,
        "Exclusivos A",
        resultado.exclusivos_a
    )

    escrever_dataframe(
        writer,
        "Exclusivos B",
        resultado.exclusivos_b
    )

    escrever_dataframe(
        writer,
        "Inconsistentes",
        resultado.inconsistentes,
        destacar_colunas=[
            "Origem",
            "Motivo da Inconsistência"
        ]
    )

    escrever_dataframe(
        writer,
        "Consolidado",
        resultado.consolidado
    )

    extras = (
        ("Duplicados A", "duplicados_a"),
        ("Duplicados B", "duplicados_b"),
        ("Sem chave A", "sem_chave_a"),
        ("Sem chave B", "sem_chave_b"),
    )

    for nome_aba, atributo in extras:
        if hasattr(resultado, atributo):
            escrever_dataframe(writer, nome_aba, getattr(resultado, atributo))

# ============================================================
# GERAÇÃO DO EXCEL
# ============================================================

def gerar_excel(resultado):

    buffer = BytesIO()

    with pd.ExcelWriter(
        buffer,
        engine="openpyxl"
    ) as writer:

        escrever_aba_dashboard(
            writer,
            resultado
        )

        escrever_planilhas(
            writer,
            resultado
        )

    buffer.seek(0)

    return buffer
